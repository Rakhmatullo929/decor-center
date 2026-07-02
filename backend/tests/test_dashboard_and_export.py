import datetime
from io import BytesIO

import pytest
from django.utils import timezone
from openpyxl import load_workbook

from apps.assessments.models import TestSession
from apps.core.excel import XLSX_CONTENT_TYPE
from apps.medical.models import MedicalCheck

from .factories import EmployeeFactory

pytestmark = pytest.mark.django_db

DASHBOARD_URL = "/api/v1/dashboard/stats/"
SESSIONS_EXPORT_URL = "/api/v1/test-sessions/export/"
CHECKS_URL = "/api/v1/medical-checks/"
CHECKS_EXPORT_URL = "/api/v1/medical-checks/export/"


def _create_session(employee, *, passed, score):
    return TestSession.objects.create(
        employee=employee,
        module="specialty",
        specialty=employee.specialty,
        total=10,
        score=score,
        passed=passed,
        face_verified=True,
        finished_at=timezone.now(),
    )


def _check_payload(employee, **overrides):
    payload = {
        "employee": employee.id,
        "bp_systolic": 120,
        "bp_diastolic": 80,
        "pulse": 72,
        "saturation": 98,
        "alcohol_positive": False,
        "hours_worked": "8.0",
        "hours_rested": "16.0",
        "conclusion": "approved",
        "note": "",
    }
    payload.update(overrides)
    return payload


def _sheet_rows(response):
    sheet = load_workbook(BytesIO(response.content)).active
    return list(sheet.iter_rows(values_only=True))


# ---------------------------------------------------------------- dashboard


def test_dashboard_stats_counts_todays_activity(admin_client, medic_client):
    employee = EmployeeFactory()
    _create_session(employee, passed=True, score=9)
    _create_session(employee, passed=False, score=3)
    medic_client.post(CHECKS_URL, _check_payload(employee), format="json")
    medic_client.post(
        CHECKS_URL, _check_payload(employee, conclusion="rejected"), format="json"
    )

    response = admin_client.get(DASHBOARD_URL)
    assert response.status_code == 200

    tests = response.data["tests"]
    assert tests["total"] == 2
    assert tests["passed"] == 1
    assert tests["failed"] == 1
    assert tests["by_module"] == [{"module": "specialty", "total": 2, "passed": 1}]

    medical = response.data["medical"]
    assert medical == {"total": 2, "approved": 1, "rejected": 1}

    assert response.data["totals"]["active_employees"] == 1


def test_dashboard_is_admin_only(medic_client, specialist_client):
    assert medic_client.get(DASHBOARD_URL).status_code == 403
    assert specialist_client.get(DASHBOARD_URL).status_code == 403


def test_dashboard_other_date_is_empty(admin_client):
    employee = EmployeeFactory()
    _create_session(employee, passed=True, score=9)

    response = admin_client.get(DASHBOARD_URL, {"date": "2000-01-01"})
    assert response.status_code == 200
    assert response.data["tests"]["total"] == 0


# ---------------------------------------------------------------- exports


def test_sessions_export_respects_filters(admin_client):
    employee = EmployeeFactory()
    _create_session(employee, passed=True, score=9)
    _create_session(employee, passed=False, score=2)

    response = admin_client.get(SESSIONS_EXPORT_URL)
    assert response.status_code == 200
    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert "test-results-" in response["Content-Disposition"]
    assert len(_sheet_rows(response)) == 3  # header + 2 sessions

    filtered = admin_client.get(SESSIONS_EXPORT_URL, {"passed": True})
    rows = _sheet_rows(filtered)
    assert len(rows) == 2  # header + 1 passed session
    assert rows[1][7] == "Passed"


def test_sessions_export_is_admin_only(medic_client, specialist_client):
    assert medic_client.get(SESSIONS_EXPORT_URL).status_code == 403
    assert specialist_client.get(SESSIONS_EXPORT_URL).status_code == 403


def test_medical_export_respects_filters(medic_client):
    first = EmployeeFactory()
    second = EmployeeFactory()
    medic_client.post(CHECKS_URL, _check_payload(first), format="json")
    medic_client.post(CHECKS_URL, _check_payload(second, conclusion="rejected"), format="json")

    response = medic_client.get(CHECKS_EXPORT_URL)
    assert response.status_code == 200
    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert len(_sheet_rows(response)) == 3

    filtered = medic_client.get(CHECKS_EXPORT_URL, {"conclusion": "rejected"})
    rows = _sheet_rows(filtered)
    assert len(rows) == 2
    assert rows[1][1] == second.full_name


def test_medical_export_is_not_for_specialists(specialist_client):
    assert specialist_client.get(CHECKS_EXPORT_URL).status_code == 403


# ---------------------------------------------------------------- export honours the date filter


def test_sessions_export_respects_date_filter(admin_client):
    employee = EmployeeFactory()
    TestSession.objects.create(
        employee=employee,
        module="specialty",
        specialty=employee.specialty,
        total=10,
        score=9,
        passed=True,
        face_verified=True,
        started_at=timezone.make_aware(datetime.datetime(2026, 6, 11, 12, 0)),
    )
    TestSession.objects.create(
        employee=employee,
        module="specialty",
        specialty=employee.specialty,
        total=10,
        score=2,
        passed=False,
        face_verified=True,
        started_at=timezone.make_aware(datetime.datetime(2026, 6, 12, 12, 0)),
    )

    response = admin_client.get(SESSIONS_EXPORT_URL, {"date": "2026-06-11"})
    rows = _sheet_rows(response)
    assert response.status_code == 200
    assert len(rows) == 2  # header + the single 2026-06-11 session
    assert rows[1][3].startswith("2026-06-11")  # 'Started' column


def test_medical_export_respects_date_filter(medic_client):
    employee = EmployeeFactory()
    id_match = medic_client.post(CHECKS_URL, _check_payload(employee), format="json").data["id"]
    medic_client.post(CHECKS_URL, _check_payload(employee, conclusion="rejected"), format="json")
    MedicalCheck.objects.filter(pk=id_match).update(
        created_at=timezone.make_aware(datetime.datetime(2026, 6, 11, 10, 0))
    )
    MedicalCheck.objects.exclude(pk=id_match).update(
        created_at=timezone.make_aware(datetime.datetime(2026, 6, 12, 10, 0))
    )

    response = medic_client.get(CHECKS_EXPORT_URL, {"date": "2026-06-11"})
    rows = _sheet_rows(response)
    assert response.status_code == 200
    assert len(rows) == 2  # header + the single 2026-06-11 check
    assert rows[1][0].startswith("2026-06-11")  # 'Date/time' column
